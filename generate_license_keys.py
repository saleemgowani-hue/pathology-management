"""
VENDOR-ONLY TOOL — run this from your own computer, never deploy it as
part of the Streamlit Cloud app. It connects straight to your production
Neon database and inserts new license keys, then writes them to an
Excel workbook you keep for yourself and hand out one at a time.

Setup (once):
    pip install sqlalchemy psycopg2-binary openpyxl
    export DATABASE_URL="postgresql://...same string as your Streamlit secret..."

Usage:
    python generate_license_keys.py                 # 50 monthly + 50 yearly
    python generate_license_keys.py 100              # 100 of each
    python generate_license_keys.py 20 monthly       # 20 monthly only
    python generate_license_keys.py 20 yearly        # 20 yearly only
"""
import os
import sys
import secrets
import string
from datetime import datetime

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db.models import Base, LicenseKey  # noqa: E402

ALPHABET = string.ascii_uppercase + string.digits
AMBIGUOUS = set("0O1I")
ALPHABET = "".join(c for c in ALPHABET if c not in AMBIGUOUS)

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "license_keys.xlsx")


def _random_key():
    groups = ["".join(secrets.choice(ALPHABET) for _ in range(5)) for _ in range(4)]
    return "-".join(groups)


def _get_session():
    url = os.environ.get("DATABASE_URL")
    if not url:
        print("ERROR: set the DATABASE_URL environment variable first (same value as your")
        print("Streamlit secret). See the instructions at the top of this file.")
        sys.exit(1)
    engine = create_engine(url, pool_pre_ping=True)
    Base.metadata.create_all(engine)  # safe no-op if tables already exist
    return sessionmaker(bind=engine)()


def _generate(session, count, plan):
    existing = {k.key_code for k in session.query(LicenseKey).all()}
    new_keys = []
    while len(new_keys) < count:
        k = _random_key()
        if k in existing or k in new_keys:
            continue
        new_keys.append(k)
        existing.add(k)
    for k in new_keys:
        session.add(LicenseKey(key_code=k, plan=plan, is_used=False))
    session.commit()
    return new_keys


def _write_workbook(session, out_path=DEFAULT_XLSX):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="0B5394", end_color="0B5394", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=11)
    mono_font = Font(name="Consolas", size=11)
    used_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    available_fill = PatternFill(start_color="E4F7EC", end_color="E4F7EC", fill_type="solid")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["License Key", "Plan", "Validity", "Status", "Used By Tenant ID", "Activated On"]
    widths = [26, 12, 12, 14, 16, 16]

    for plan, sheet_name, validity_label in [("monthly", "Monthly Keys", "30 days"),
                                               ("yearly", "Yearly Keys", "365 days")]:
        ws = wb.create_sheet(sheet_name)
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 22

        keys = session.query(LicenseKey).filter_by(plan=plan).order_by(LicenseKey.id).all()
        for r, k in enumerate(keys, start=2):
            status = "Used" if k.is_used else "Available"
            fill = used_fill if k.is_used else available_fill
            activated = k.activated_date.strftime("%d-%b-%Y") if k.activated_date else "-"
            row_values = [k.key_code, plan.capitalize(), validity_label, status,
                          k.used_by_tenant_id or "-", activated]
            for col, val in enumerate(row_values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = mono_font if col == 1 else normal_font
                cell.alignment, cell.border, cell.fill = center, border, fill
        ws.freeze_panes = "A2"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.append(["PathoLab Pro (Cloud) — License Key Summary"])
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="0B5394")
    ws_sum.append([f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"])
    ws_sum["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws_sum.append([])
    ws_sum.append(["Plan", "Total Keys", "Used", "Available"])
    for col in range(1, 5):
        c = ws_sum.cell(row=4, column=col)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
    for i, plan in enumerate(["monthly", "yearly"], start=5):
        keys = session.query(LicenseKey).filter_by(plan=plan).all()
        used = sum(1 for k in keys if k.is_used)
        for col, val in enumerate([plan.capitalize(), len(keys), used, len(keys) - used], start=1):
            c = ws_sum.cell(row=i, column=col, value=val)
            c.font, c.alignment, c.border = normal_font, center, border
    for col, w in enumerate([20, 14, 10, 12], start=1):
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    session = _get_session()
    if len(sys.argv) >= 3 and sys.argv[2] in ("monthly", "yearly"):
        n = int(sys.argv[1])
        new_monthly = _generate(session, n, "monthly") if sys.argv[2] == "monthly" else []
        new_yearly = _generate(session, n, "yearly") if sys.argv[2] == "yearly" else []
    else:
        n = int(sys.argv[1]) if len(sys.argv) > 1 else 50
        new_monthly = _generate(session, n, "monthly")
        new_yearly = _generate(session, n, "yearly")

    out_path = _write_workbook(session)
    print(f"Generated {len(new_monthly)} monthly + {len(new_yearly)} yearly keys.")
    print(f"Workbook written to: {out_path}")
